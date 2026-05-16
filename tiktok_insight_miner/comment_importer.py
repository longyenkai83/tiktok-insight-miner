"""Manual comment importer — đọc CSV/XLSX → chuẩn hoá thành raw_comments.json.

Cho phép nạp comment từ Facebook, YouTube, fanpage, group, hoặc bất kỳ nguồn nào
mà user copy-paste được vào CSV/Excel — không phụ thuộc TikTok scraper.

Schema CSV/XLSX:
  required: 'comment' HOẶC 'text'
  optional: 'platform', 'source_url', 'author', 'likes', 'replies', 'created_at'

Output schema khớp 100% với scraper.py để pipeline cũ (classify/bank/select/production)
chạy được ngay không phải sửa gì.
"""

from __future__ import annotations

import csv
import logging
from datetime import date
from pathlib import Path
from typing import Any, Iterator

from tiktok_insight_miner.models import Comment
from tiktok_insight_miner.scraper import save_comments_json

logger = logging.getLogger(__name__)

# Tên cột chấp nhận cho text (ưu tiên 'comment' nếu có cả 2)
TEXT_COLUMN_CANDIDATES = ("comment", "text")

# Default values khi cột thiếu hoặc empty
DEFAULTS = {
    "platform": "manual",
    "source_url": "",
    "author": "unknown",
    "likes": 0,
    "replies": 0,
    "created_at": "",
}

MIN_TEXT_LENGTH = 5


# --- Reader: dispatch theo extension ---

def read_rows(path: Path) -> Iterator[dict[str, Any]]:
    """Đọc CSV hoặc XLSX → yield dict mỗi row.

    Auto-detect:
    - .csv: native csv module, support UTF-8 và UTF-8-BOM (Excel-saved)
    - .xlsx: openpyxl (lazy import — fail rõ nếu chưa cài)
    """
    ext = path.suffix.lower()
    if ext == ".csv":
        yield from _read_csv(path)
    elif ext in (".xlsx", ".xlsm"):
        yield from _read_xlsx(path)
    else:
        raise ValueError(
            f"Format không hỗ trợ: {ext}. Chỉ chấp nhận .csv hoặc .xlsx"
        )


def _read_csv(path: Path) -> Iterator[dict[str, Any]]:
    # utf-8-sig tự strip BOM nếu file save từ Excel
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield {k.strip().lower() if k else k: v for k, v in row.items()}


def _read_xlsx(path: Path) -> Iterator[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "Đọc Excel cần `openpyxl`. Cài thêm: `pip install openpyxl`. "
            "Hoặc save file thành .csv (Excel: File → Save As → CSV UTF-8)."
        ) from e

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    for row in rows_iter:
        if not any(c is not None and str(c).strip() for c in row):
            continue  # bỏ row hoàn toàn trống
        yield {h: row[i] if i < len(row) else None for i, h in enumerate(headers)}

    wb.close()


# --- Validate header ---

def detect_text_column(headers: set[str]) -> str:
    """Trả tên cột chứa text comment, hoặc raise nếu không có."""
    for candidate in TEXT_COLUMN_CANDIDATES:
        if candidate in headers:
            return candidate
    raise ValueError(
        f"CSV/XLSX phải có 1 trong các cột: {TEXT_COLUMN_CANDIDATES}. "
        f"Thấy headers: {sorted(headers)}"
    )


# --- Row → Comment conversion ---

def _safe_int(v: Any, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return default


def _safe_str(v: Any, default: str = "") -> str:
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def row_to_comment(
    row: dict[str, Any],
    idx: int,
    text_column: str,
    source_override: str | None = None,
) -> Comment | None:
    """Convert 1 row → Comment object. Trả None nếu skip (empty/short).

    Args:
        row: dict đã lowercase keys
        idx: 1-based row number, dùng cho id
        text_column: 'comment' hoặc 'text'
        source_override: nếu set, override cột 'platform' empty
    """
    text = _safe_str(row.get(text_column))
    if not text:
        return None  # skip empty
    if len(text) < MIN_TEXT_LENGTH:
        return None  # skip too-short

    # Platform: CSV col > flag override > default
    platform = _safe_str(row.get("platform"))
    if not platform:
        platform = source_override or DEFAULTS["platform"]

    return Comment(
        id=f"manual-{idx:04d}",
        text=text,
        author=_safe_str(row.get("author"), DEFAULTS["author"]),
        likes=_safe_int(row.get("likes"), DEFAULTS["likes"]),
        reply_count=_safe_int(row.get("replies"), DEFAULTS["replies"]),
        created_at=_safe_str(row.get("created_at"), DEFAULTS["created_at"]),
        video_url=_safe_str(row.get("source_url"), DEFAULTS["source_url"]),
        raw={
            "platform": platform,
            "imported_from": "manual_csv",
            "original_row": {k: (str(v) if v is not None else "") for k, v in row.items()},
        },
    )


# --- Output path ---

def resolve_output_path(
    niche_slug: str,
    output_root: Path,
    run_date: str | None = None,
) -> Path:
    """Trả `output/<niche>/<date>__manual-import/raw_comments.json`."""
    run_date = run_date or date.today().isoformat()
    folder = output_root / niche_slug / f"{run_date}__manual-import"
    folder.mkdir(parents=True, exist_ok=True)
    return folder / "raw_comments.json"


# --- Entry point ---

def import_comments(
    input_path: Path,
    niche_slug: str,
    output_root: Path,
    source_override: str | None = None,
    run_date: str | None = None,
) -> dict[str, Any]:
    """All-in-one: read CSV/XLSX → validate → convert → save raw_comments.json.

    Returns stats dict: total_rows, kept, skipped_empty, skipped_short, output_path.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"File input không tồn tại: {input_path}")

    # Read all rows first để biết headers
    all_rows = list(read_rows(input_path))
    if not all_rows:
        raise ValueError(f"File {input_path} không có data")

    headers = set()
    for r in all_rows:
        headers.update(r.keys())

    text_col = detect_text_column(headers)
    logger.info("Phát hiện cột text: '%s'", text_col)

    stats = {
        "total_rows": len(all_rows),
        "kept": 0,
        "skipped_empty": 0,
        "skipped_short": 0,
    }
    comments: list[Comment] = []

    for idx, row in enumerate(all_rows, 1):
        text = _safe_str(row.get(text_col))
        if not text:
            stats["skipped_empty"] += 1
            continue
        if len(text) < MIN_TEXT_LENGTH:
            stats["skipped_short"] += 1
            logger.debug("Row %d: skip (text too short: %r)", idx, text)
            continue

        comment = row_to_comment(row, idx, text_col, source_override=source_override)
        if comment is None:
            stats["skipped_empty"] += 1
            continue
        comments.append(comment)
        stats["kept"] += 1

    output_path = resolve_output_path(niche_slug, output_root, run_date=run_date)
    save_comments_json(comments, output_path)
    logger.info(
        "Imported %d comments → %s (skip %d empty, %d short)",
        stats["kept"], output_path, stats["skipped_empty"], stats["skipped_short"],
    )

    stats["output_path"] = output_path
    stats["text_column"] = text_col
    stats["niche_slug"] = niche_slug
    return stats
